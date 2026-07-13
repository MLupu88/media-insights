"""Builds the Excel export (Phase 6C) from a `ProjectReportData`/
`ComparisonReportData` instance (`app/services/report_data.py`).

One function per sheet; two orchestrators (`build_project_xlsx`,
`build_comparison_xlsx`). Every cell is a pre-computed literal value — no
formulas anywhere, so the workbook can never recompute a different number
than the backend already did. Ranked sheets are capped at `EXCEL_TOP_N`
(currently an identity slice against the data layer's own `MAX_TOP_N`
fetch, written explicitly so the contract stays correct if that ever
changes); Article Detail and labels are never truncated here — only PPTX
truncates for slide legibility.
"""

from io import BytesIO

from openpyxl import Workbook

from app.services import report_xlsx_style as style
from app.services.report_contract import EXCEL_TOP_N, MAX_XLSX_BYTES, NO_DATA_MESSAGE
from app.services.report_data import ComparisonReportData, ProjectReportData, ReportTooLargeError


def _metadata_pairs(metadata, extra: list[tuple[str, str]] | None = None) -> list[tuple[str, str]]:
    coverage = metadata.article_detail_coverage
    insight_coverage = metadata.insight_coverage
    pairs = [
        ("Scope", metadata.scope_label),
        ("Filters", metadata.filters_label),
        ("Generated", metadata.generated_at.strftime("%Y-%m-%d %H:%M UTC")),
        ("Population", metadata.population_definition),
        ("Chat exclusion", metadata.chat_exclusion_note),
        (
            "Article detail coverage",
            f"Showing {coverage.shown_count:,} of {coverage.total_count:,} eligible articles"
            + (" (truncated)" if coverage.truncated else ""),
        ),
        (
            "Insight coverage",
            f"{insight_coverage.available_count} valid insight(s) available, "
            f"{insight_coverage.included_count} included, "
            f"{insight_coverage.excluded_causal_count} excluded (unsupported causal language)",
        ),
    ]
    if extra:
        pairs.extend(extra)
    return pairs


def _write_no_data(ws) -> None:
    ws.cell(row=1, column=1, value=NO_DATA_MESSAGE).font = style.LABEL_FONT


def _check_size(wb: Workbook) -> bytes:
    buffer = BytesIO()
    wb.save(buffer)
    data = buffer.getvalue()
    if len(data) > MAX_XLSX_BYTES:
        raise ReportTooLargeError(
            f"Generated Excel workbook exceeded the maximum supported size "
            f"({MAX_XLSX_BYTES:,} bytes)."
        )
    return data


# --- shared sheet builders (used by both workbooks) -----------------------------


def _sheet_kpi_summary(wb: Workbook, title: str, kpis: dict) -> None:
    ws = wb.create_sheet(style.safe_sheet_name(title))
    headers = ["Metric", "Value", "Notes"]
    style.write_header_row(ws, 1, headers)
    rows = [
        ("Total imported rows", kpis["total_imported_rows"], "Pipeline total, unaffected by filters"),
        ("Valid rows", kpis["valid_rows"], "Pipeline total"),
        ("Invalid rows", kpis["invalid_rows"], "Pipeline total"),
        ("Duplicate rows", kpis["duplicate_rows"], "Pipeline total"),
        ("Duplicate share", kpis["duplicate_share_pct"], "% of valid rows, pipeline total"),
        ("Unique valid articles", kpis["unique_valid_articles"], "Filtered analytical population"),
        ("Unique classified articles", kpis["unique_classified_articles"], ""),
        ("Unique unclassified articles", kpis["unique_unclassified_articles"], ""),
        ("Total reach", kpis["total_reach"], ""),
        ("Average reach", kpis["average_reach"], ""),
        ("Median reach", kpis["median_reach"], ""),
        ("Publications", kpis["publication_count"], ""),
        ("Low-confidence classifications", kpis["low_confidence_count"], ""),
    ]
    for i, (label, value, note) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=label)
        if "share" in label.lower():
            style.set_percent(ws.cell(row=i, column=2), value)
        else:
            style.set_number(ws.cell(row=i, column=2), value, decimals=1)
        ws.cell(row=i, column=3, value=note)
    style.freeze_header(ws)
    style.apply_autofilter(ws, 1, len(headers), len(rows))
    style.set_column_widths(ws, [32, 16, 46])


def _sheet_brand_performance(wb: Workbook, title: str, brand_rows: list[dict]) -> None:
    ws = wb.create_sheet(style.safe_sheet_name(title))
    if not brand_rows:
        _write_no_data(ws)
        return
    headers = [
        "Brand", "Article Count", "SOV %", "Total Reach", "Reach Share %",
        "Avg Reach", "Median Reach", "Primary Focus Count", "Mentioned Only Count",
    ]
    style.write_header_row(ws, 1, headers)
    rows = brand_rows[:EXCEL_TOP_N]
    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=r["brand"])
        style.set_number(ws.cell(row=i, column=2), r["article_count"])
        style.set_percent(ws.cell(row=i, column=3), r["sov_pct"])
        style.set_number(ws.cell(row=i, column=4), r["total_reach"], decimals=1)
        style.set_percent(ws.cell(row=i, column=5), r["reach_share_pct"])
        style.set_number(ws.cell(row=i, column=6), r["average_reach"], decimals=1)
        style.set_number(ws.cell(row=i, column=7), r["median_reach"], decimals=1)
        style.set_number(ws.cell(row=i, column=8), r["primary_focus_count"])
        style.set_number(ws.cell(row=i, column=9), r["mentioned_only_count"])
    style.freeze_header(ws)
    style.apply_autofilter(ws, 1, len(headers), len(rows))
    style.set_column_widths(ws, [24, 14, 10, 14, 14, 12, 12, 18, 18])


def _sheet_topic_distribution(wb: Workbook, title: str, topics: dict) -> None:
    ws = wb.create_sheet(style.safe_sheet_name(title))
    primary = topics["primary_topic_distribution"][:EXCEL_TOP_N]
    if not primary:
        _write_no_data(ws)
        return
    row = 1
    ws.cell(row=row, column=1, value="Primary Topic Distribution").font = style.LABEL_FONT
    row += 1
    headers = ["Primary Topic", "Count", "%", "Total Reach"]
    style.write_header_row(ws, row, headers)
    header_row = row
    row += 1
    for r in primary:
        ws.cell(row=row, column=1, value=r["value"])
        style.set_number(ws.cell(row=row, column=2), r["count"])
        style.set_percent(ws.cell(row=row, column=3), r["pct"])
        style.set_number(ws.cell(row=row, column=4), r["total_reach"], decimals=1)
        row += 1
    style.apply_autofilter(ws, header_row, len(headers), len(primary))
    style.freeze_header(ws, header_row)

    row += 1
    ws.cell(row=row, column=1, value="Communication Category Distribution").font = style.LABEL_FONT
    row += 1
    style.write_header_row(ws, row, headers)
    row += 1
    for r in topics["communication_category_distribution"][:EXCEL_TOP_N]:
        ws.cell(row=row, column=1, value=r["value"])
        style.set_number(ws.cell(row=row, column=2), r["count"])
        style.set_percent(ws.cell(row=row, column=3), r["pct"])
        style.set_number(ws.cell(row=row, column=4), r["total_reach"], decimals=1)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Topic Mix by Brand").font = style.LABEL_FONT
    row += 1
    ws.cell(row=row, column=1, value="Brand")
    ws.cell(row=row, column=2, value="Topics (topic: count, %)")
    row += 1
    for brand_row in topics["topic_mix_by_brand"]:
        ws.cell(row=row, column=1, value=brand_row["brand"])
        summary = "; ".join(f"{t['topic']}: {t['count']} ({t['pct']}%)" for t in brand_row["topics"])
        ws.cell(row=row, column=2, value=summary)
        row += 1

    style.set_column_widths(ws, [30, 14, 10, 16])


def _sheet_sentiment(wb: Workbook, title: str, sentiment: dict) -> None:
    ws = wb.create_sheet(style.safe_sheet_name(title))
    overall = sentiment["overall_distribution"]
    if not overall:
        _write_no_data(ws)
        return
    row = 1
    ws.cell(row=row, column=1, value="Overall Sentiment").font = style.LABEL_FONT
    row += 1
    headers = ["Sentiment", "Count", "%"]
    style.write_header_row(ws, row, headers)
    row += 1
    for r in overall:
        ws.cell(row=row, column=1, value=r["value"])
        style.set_number(ws.cell(row=row, column=2), r["count"])
        style.set_percent(ws.cell(row=row, column=3), r["pct"])
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Brand Role").font = style.LABEL_FONT
    row += 1
    style.write_header_row(ws, row, ["Role", "Count", "%"])
    row += 1
    for r in sentiment["brand_role_distribution"]:
        ws.cell(row=row, column=1, value=r["value"])
        style.set_number(ws.cell(row=row, column=2), r["count"])
        style.set_percent(ws.cell(row=row, column=3), r["pct"])
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Sentiment by Brand").font = style.LABEL_FONT
    row += 1
    style.write_header_row(ws, row, ["Brand", "Total", "Positive", "Neutral", "Negative", "Mixed"])
    row += 1
    for r in sentiment["sentiment_by_brand"]:
        ws.cell(row=row, column=1, value=r["brand"])
        style.set_number(ws.cell(row=row, column=2), r["total"])
        style.set_number(ws.cell(row=row, column=3), r["counts"].get("positive", 0))
        style.set_number(ws.cell(row=row, column=4), r["counts"].get("neutral", 0))
        style.set_number(ws.cell(row=row, column=5), r["counts"].get("negative", 0))
        style.set_number(ws.cell(row=row, column=6), r["counts"].get("mixed", 0))
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Low-Confidence Items").font = style.LABEL_FONT
    row += 1
    style.write_header_row(ws, row, ["Title", "Brand", "Primary Topic", "Confidence"])
    row += 1
    for item in sentiment["low_confidence_items"]:
        ws.cell(row=row, column=1, value=item.get("title"))
        ws.cell(row=row, column=2, value=item.get("brand"))
        ws.cell(row=row, column=3, value=item.get("primary_topic"))
        style.set_number(ws.cell(row=row, column=4), item.get("confidence"), decimals=2)
        row += 1

    style.set_column_widths(ws, [30, 14, 14, 12, 12, 12])


def _sheet_publications(wb: Workbook, title: str, pubs: dict) -> None:
    ws = wb.create_sheet(style.safe_sheet_name(title))
    rows = pubs["publications_by_volume"][:EXCEL_TOP_N]
    if not rows:
        _write_no_data(ws)
        return
    headers = ["Publication", "Article Count", "Volume %", "Total Reach", "Reach %"]
    style.write_header_row(ws, 1, headers)
    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=r["publication"])
        style.set_number(ws.cell(row=i, column=2), r["article_count"])
        style.set_percent(ws.cell(row=i, column=3), r["volume_pct"])
        style.set_number(ws.cell(row=i, column=4), r["total_reach"], decimals=1)
        style.set_percent(ws.cell(row=i, column=5), r["reach_pct"])
    style.freeze_header(ws)
    style.apply_autofilter(ws, 1, len(headers), len(rows))

    conc_row = len(rows) + 3
    ws.cell(row=conc_row, column=1, value="Concentration").font = style.LABEL_FONT
    conc = pubs["publication_concentration"]
    for offset, (label, key) in enumerate((
        ("Top 3 by volume", "top3_volume_pct"), ("Top 5 by volume", "top5_volume_pct"),
        ("Top 3 by reach", "top3_reach_pct"), ("Top 5 by reach", "top5_reach_pct"),
    )):
        ws.cell(row=conc_row + 1 + offset, column=1, value=label)
        style.set_percent(ws.cell(row=conc_row + 1 + offset, column=2), conc[key])

    style.set_column_widths(ws, [34, 14, 12, 14, 12])


def _sheet_stories(wb: Workbook, title: str, pubs: dict) -> None:
    ws = wb.create_sheet(style.safe_sheet_name(title))
    rows = pubs["stories_by_volume"][:EXCEL_TOP_N]
    if not rows:
        _write_no_data(ws)
        return
    headers = ["Story Key", "Article Count", "Total Reach"]
    style.write_header_row(ws, 1, headers)
    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=r["story_key"])
        style.set_number(ws.cell(row=i, column=2), r["article_count"])
        style.set_number(ws.cell(row=i, column=3), r["total_reach"], decimals=1)
    style.freeze_header(ws)
    style.apply_autofilter(ws, 1, len(headers), len(rows))

    conc_row = len(rows) + 3
    ws.cell(row=conc_row, column=1, value="Concentration").font = style.LABEL_FONT
    conc = pubs["story_concentration"]
    for offset, (label, key) in enumerate((
        ("Top 3 by volume", "top3_volume_pct"), ("Top 5 by volume", "top5_volume_pct"),
        ("Top 3 by reach", "top3_reach_pct"), ("Top 5 by reach", "top5_reach_pct"),
    )):
        ws.cell(row=conc_row + 1 + offset, column=1, value=label)
        style.set_percent(ws.cell(row=conc_row + 1 + offset, column=2), conc[key])

    style.set_column_widths(ws, [34, 14, 14])


def _sheet_validated_insights(wb: Workbook, title: str, insights: list) -> None:
    ws = wb.create_sheet(style.safe_sheet_name(title))
    if not insights:
        _write_no_data(ws)
        return
    headers = [
        "Type", "Narrative Type", "Title", "Narrative", "Related Brand",
        "Related Topic", "Related Publication", "Related Story", "Confidence", "Caveat",
    ]
    style.write_header_row(ws, 1, headers)
    for i, insight in enumerate(insights, start=2):
        ws.cell(row=i, column=1, value=insight.label)
        ws.cell(row=i, column=2, value=insight.narrative_type)
        ws.cell(row=i, column=3, value=insight.title)
        ws.cell(row=i, column=4, value=insight.narrative)
        ws.cell(row=i, column=5, value=insight.related_brand)
        ws.cell(row=i, column=6, value=insight.related_topic)
        ws.cell(row=i, column=7, value=insight.related_publication)
        ws.cell(row=i, column=8, value=insight.related_story_key)
        style.set_number(ws.cell(row=i, column=9), insight.confidence, decimals=2)
        ws.cell(row=i, column=10, value=insight.caveat)
    style.freeze_header(ws)
    style.apply_autofilter(ws, 1, len(headers), len(insights))
    style.set_column_widths(ws, [16, 24, 30, 70, 18, 18, 22, 22, 12, 40])


def _sheet_article_detail(wb: Workbook, title: str, article_detail: list, with_period: bool) -> None:
    ws = wb.create_sheet(style.safe_sheet_name(title))
    if not article_detail:
        _write_no_data(ws)
        return
    headers = ["Article ID", "Title", "Brand", "Publication", "Date", "Primary Topic",
               "Communication Category", "Sentiment", "Brand Role", "Confidence", "Reach",
               "Article URL", "Mediatrust URL"]
    if with_period:
        headers = ["Period"] + headers
    style.write_header_row(ws, 1, headers)
    for i, row in enumerate(article_detail, start=2):
        col = 1
        if with_period:
            ws.cell(row=i, column=col, value=row.period)
            col += 1
        ws.cell(row=i, column=col, value=str(row.article_id)); col += 1
        ws.cell(row=i, column=col, value=row.title); col += 1
        ws.cell(row=i, column=col, value=row.brand); col += 1
        ws.cell(row=i, column=col, value=row.publication); col += 1
        ws.cell(row=i, column=col, value=row.publication_date.isoformat() if row.publication_date else None); col += 1
        ws.cell(row=i, column=col, value=row.primary_topic); col += 1
        ws.cell(row=i, column=col, value=row.communication_category); col += 1
        ws.cell(row=i, column=col, value=row.sentiment); col += 1
        ws.cell(row=i, column=col, value=row.brand_role); col += 1
        style.set_number(ws.cell(row=i, column=col), row.confidence, decimals=2); col += 1
        style.set_number(ws.cell(row=i, column=col), row.reach, decimals=1); col += 1
        ws.cell(row=i, column=col, value=row.article_url); col += 1
        ws.cell(row=i, column=col, value=row.mediatrust_url); col += 1
    style.freeze_header(ws)
    style.apply_autofilter(ws, 1, len(headers), len(article_detail))
    widths = [30, 40, 18, 26, 12, 20, 22, 12, 16, 12, 12, 40, 40]
    if with_period:
        widths = [12] + widths
    style.set_column_widths(ws, widths)


def _sheet_methodology(wb: Workbook, title: str, metadata) -> None:
    ws = wb.create_sheet(style.safe_sheet_name(title))
    header_row = style.add_branded_header(ws, "MSL The Practice — Methodology & Filters")
    style.write_metadata_block(ws, _metadata_pairs(metadata), start_row=header_row)
    style.set_column_widths(ws, [26, 90])


# --- project workbook -------------------------------------------------------


def build_project_xlsx(data: ProjectReportData) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Executive Summary"
    header_row = style.add_branded_header(ws, f"MSL The Practice — {data.project_name}")
    row = style.write_metadata_block(ws, _metadata_pairs(data.metadata), start_row=header_row)
    kpis = data.analytics["kpis"]
    style.write_header_row(ws, row, ["Metric", "Value"])
    kpi_row = row + 1
    for label, value in (
        ("Unique valid articles", kpis["unique_valid_articles"]),
        ("Total reach", kpis["total_reach"]),
        ("Average reach", kpis["average_reach"]),
        ("Publications", kpis["publication_count"]),
    ):
        ws.cell(row=kpi_row, column=1, value=label)
        style.set_number(ws.cell(row=kpi_row, column=2), value, decimals=1)
        kpi_row += 1
    insight = next((i for i in data.insights if i.narrative_type == "executive_summary"), None)
    if insight:
        ws.cell(row=kpi_row + 1, column=1, value=f"{insight.label}:").font = style.LABEL_FONT
        ws.cell(row=kpi_row + 2, column=1, value=insight.narrative)
    style.set_column_widths(ws, [30, 70])

    _sheet_kpi_summary(wb, "KPI Summary", kpis)
    _sheet_brand_performance(wb, "Brand Performance", data.analytics["brands"]["by_volume"])
    _sheet_topic_distribution(wb, "Topic Distribution", data.analytics["topics"])
    _sheet_sentiment(wb, "Sentiment", data.analytics["sentiment"])
    _sheet_publications(wb, "Publications", data.analytics["publications_and_stories"])
    _sheet_stories(wb, "Story Clusters", data.analytics["publications_and_stories"])
    _sheet_validated_insights(wb, "Validated Insights", data.insights)
    _sheet_article_detail(wb, "Article Detail", data.article_detail, with_period=False)
    _sheet_methodology(wb, "Methodology", data.metadata)

    return _check_size(wb)


# --- comparison workbook -----------------------------------------------------


def _sheet_kpi_comparison(wb: Workbook, kpi_deltas: dict) -> None:
    ws = wb.create_sheet("KPI Comparison")
    headers = ["Metric", "Baseline", "Comparison", "Absolute Delta", "% Delta"]
    style.write_header_row(ws, 1, headers)
    row = 2
    for key, d in kpi_deltas.items():
        ws.cell(row=row, column=1, value=key.replace("_", " ").title())
        style.set_number(ws.cell(row=row, column=2), d["baseline"], decimals=1)
        style.set_number(ws.cell(row=row, column=3), d["comparison"], decimals=1)
        style.set_number(ws.cell(row=row, column=4), d["absolute_delta"], decimals=1)
        if d["percentage_delta"] is not None:
            style.set_percent(ws.cell(row=row, column=5), d["percentage_delta"])
        else:
            ws.cell(row=row, column=5, value=None)
        row += 1
    style.freeze_header(ws)
    style.apply_autofilter(ws, 1, len(headers), len(kpi_deltas))
    style.set_column_widths(ws, [32, 16, 16, 16, 12])


def _sheet_brand_comparison(wb: Workbook, brand_deltas: list[dict]) -> None:
    ws = wb.create_sheet("Brand Performance Comparison")
    if not brand_deltas:
        _write_no_data(ws)
        return
    headers = [
        "Brand", "Baseline SOV %", "Comparison SOV %", "SOV Delta (pp)",
        "Baseline Reach Share %", "Comparison Reach Share %", "Reach Share Delta (pp)",
        "Baseline Rank", "Comparison Rank", "Rank Change", "New Entrant", "Dropout",
    ]
    style.write_header_row(ws, 1, headers)
    rows = brand_deltas[:EXCEL_TOP_N]
    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=r["brand"])
        style.set_percent(ws.cell(row=i, column=2), r["baseline_sov_pct"])
        style.set_percent(ws.cell(row=i, column=3), r["comparison_sov_pct"])
        style.set_pp(ws.cell(row=i, column=4), r["sov_delta_pp"])
        style.set_percent(ws.cell(row=i, column=5), r["baseline_reach_share_pct"])
        style.set_percent(ws.cell(row=i, column=6), r["comparison_reach_share_pct"])
        style.set_pp(ws.cell(row=i, column=7), r["reach_share_delta_pp"])
        style.set_number(ws.cell(row=i, column=8), r["baseline_rank"])
        style.set_number(ws.cell(row=i, column=9), r["comparison_rank"])
        style.set_number(ws.cell(row=i, column=10), r["rank_change"])
        ws.cell(row=i, column=11, value="Yes" if r["is_new_entrant"] else "No")
        ws.cell(row=i, column=12, value="Yes" if r["is_dropout"] else "No")
    style.freeze_header(ws)
    style.apply_autofilter(ws, 1, len(headers), len(rows))
    style.set_column_widths(ws, [22, 16, 18, 14, 18, 20, 18, 14, 16, 12, 12, 10])


def _sheet_distribution_comparison(wb: Workbook, title: str, section_label: str, dist: dict) -> None:
    ws = wb.create_sheet(style.safe_sheet_name(title))
    rows = dist["rows"]
    if not rows:
        _write_no_data(ws)
        return
    headers = [section_label, "Baseline %", "Comparison %", "Delta (pp)", "New", "Gone"]
    style.write_header_row(ws, 1, headers)
    for i, r in enumerate(rows[:EXCEL_TOP_N], start=2):
        ws.cell(row=i, column=1, value=r["value"])
        style.set_percent(ws.cell(row=i, column=2), r["baseline_pct"])
        style.set_percent(ws.cell(row=i, column=3), r["comparison_pct"])
        style.set_pp(ws.cell(row=i, column=4), r["pct_delta_pp"])
        ws.cell(row=i, column=5, value="Yes" if r["is_new"] else "No")
        ws.cell(row=i, column=6, value="Yes" if r["is_gone"] else "No")
    style.freeze_header(ws)
    style.apply_autofilter(ws, 1, len(headers), min(len(rows), EXCEL_TOP_N))
    style.set_column_widths(ws, [28, 14, 14, 12, 8, 8])


def _sheet_ranking_movement(wb: Workbook, title: str, key_label: str, key_field: str, rows: list[dict]) -> None:
    ws = wb.create_sheet(style.safe_sheet_name(title))
    if not rows:
        _write_no_data(ws)
        return
    headers = [key_label, "Baseline Rank", "Comparison Rank", "Rank Change", "New Entrant", "Dropout"]
    style.write_header_row(ws, 1, headers)
    top_rows = rows[:EXCEL_TOP_N]
    for i, r in enumerate(top_rows, start=2):
        ws.cell(row=i, column=1, value=r[key_field])
        style.set_number(ws.cell(row=i, column=2), r["baseline_rank"])
        style.set_number(ws.cell(row=i, column=3), r["comparison_rank"])
        style.set_number(ws.cell(row=i, column=4), r["rank_change"])
        ws.cell(row=i, column=5, value="Yes" if r["is_new_entrant"] else "No")
        ws.cell(row=i, column=6, value="Yes" if r["is_dropout"] else "No")
    style.freeze_header(ws)
    style.apply_autofilter(ws, 1, len(headers), len(top_rows))
    style.set_column_widths(ws, [32, 14, 16, 12, 12, 10])


def _sheet_concentration_volatility(wb: Workbook, comparison: dict) -> None:
    ws = wb.create_sheet("Concentration & Volatility")
    row = 1
    ws.cell(row=row, column=1, value="Publication Concentration Delta").font = style.LABEL_FONT
    row += 1
    style.write_header_row(ws, row, ["Metric", "Baseline", "Comparison", "Delta (pp)"])
    row += 1
    for label, key in (
        ("Top 3 by volume", "top3_volume_pct"), ("Top 5 by volume", "top5_volume_pct"),
        ("Top 3 by reach", "top3_reach_pct"), ("Top 5 by reach", "top5_reach_pct"),
    ):
        item = comparison["deltas"]["publication_concentration"][key]
        ws.cell(row=row, column=1, value=label)
        style.set_percent(ws.cell(row=row, column=2), item["baseline"])
        style.set_percent(ws.cell(row=row, column=3), item["comparison"])
        style.set_pp(ws.cell(row=row, column=4), item["delta_pp"])
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Story Concentration Delta").font = style.LABEL_FONT
    row += 1
    style.write_header_row(ws, row, ["Metric", "Baseline", "Comparison", "Delta (pp)"])
    row += 1
    for label, key in (
        ("Top 3 by volume", "top3_volume_pct"), ("Top 5 by volume", "top5_volume_pct"),
        ("Top 3 by reach", "top3_reach_pct"), ("Top 5 by reach", "top5_reach_pct"),
    ):
        item = comparison["deltas"]["story_concentration"][key]
        ws.cell(row=row, column=1, value=label)
        style.set_percent(ws.cell(row=row, column=2), item["baseline"])
        style.set_percent(ws.cell(row=row, column=3), item["comparison"])
        style.set_pp(ws.cell(row=row, column=4), item["delta_pp"])
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Volatility").font = style.LABEL_FONT
    row += 1
    volatility = comparison["volatility"]
    for label, value in (
        ("Publications: avg rank change", volatility["publications"]["avg_rank_change"]),
        ("Publications: new entrants", volatility["publications"]["entrants_count"]),
        ("Publications: dropouts", volatility["publications"]["dropouts_count"]),
        ("Stories: avg rank change", volatility["stories"]["avg_rank_change"]),
        ("Stories: new entrants", volatility["stories"]["entrants_count"]),
        ("Stories: dropouts", volatility["stories"]["dropouts_count"]),
    ):
        ws.cell(row=row, column=1, value=label)
        style.set_number(ws.cell(row=row, column=2), value, decimals=1)
        row += 1

    style.set_column_widths(ws, [32, 14, 14, 12])


def build_comparison_xlsx(data: ComparisonReportData) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Executive Summary"
    header_row = style.add_branded_header(
        ws, f"MSL The Practice — {data.baseline_label} vs {data.comparison_label}"
    )
    row = style.write_metadata_block(ws, _metadata_pairs(data.metadata), start_row=header_row)
    kpi_deltas = data.comparison["deltas"]["kpis"]
    style.write_header_row(ws, row, ["Metric", "Baseline", "Comparison", "Delta"])
    kpi_row = row + 1
    for key, label in (
        ("unique_valid_articles", "Unique valid articles"),
        ("total_reach", "Total reach"),
        ("publication_count", "Publications"),
    ):
        d = kpi_deltas[key]
        ws.cell(row=kpi_row, column=1, value=label)
        style.set_number(ws.cell(row=kpi_row, column=2), d["baseline"], decimals=1)
        style.set_number(ws.cell(row=kpi_row, column=3), d["comparison"], decimals=1)
        style.set_number(ws.cell(row=kpi_row, column=4), d["absolute_delta"], decimals=1)
        kpi_row += 1
    insight = next(
        (i for i in data.insights if i.narrative_type == "comparison_executive_summary"), None
    )
    if insight:
        ws.cell(row=kpi_row + 1, column=1, value=f"{insight.label}:").font = style.LABEL_FONT
        ws.cell(row=kpi_row + 2, column=1, value=insight.narrative)
    style.set_column_widths(ws, [30, 18, 18, 70])

    _sheet_kpi_comparison(wb, kpi_deltas)
    _sheet_brand_comparison(wb, data.comparison["deltas"]["brands"])
    _sheet_distribution_comparison(wb, "Topic & Category Shifts", "Topic", data.comparison["deltas"]["topics"])
    _sheet_distribution_comparison(wb, "Sentiment & Brand Role", "Sentiment", data.comparison["deltas"]["sentiment"])
    _sheet_ranking_movement(
        wb, "Publication Movement", "Publication", "publication",
        data.comparison["deltas"]["publications_by_volume"],
    )
    _sheet_ranking_movement(
        wb, "Story Movement", "Story Key", "story_key", data.comparison["deltas"]["stories_by_volume"]
    )
    _sheet_concentration_volatility(wb, data.comparison)
    _sheet_validated_insights(wb, "Validated Insights", data.insights)
    _sheet_article_detail(wb, "Article Detail", data.article_detail, with_period=True)
    _sheet_methodology(wb, "Methodology", data.metadata)

    return _check_size(wb)
