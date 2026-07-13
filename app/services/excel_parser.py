import hashlib
import re
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.models.article import ImportStatus, RetailerReviewStatus
from app.services.retailers import assign_retailer, compute_dominant_retailer

MIN_HEADER_MATCHES = 4
MAX_HEADER_SCAN_ROWS = 15

RETAILER_COLUMN_FIELD = "retailer_column"

_RAW_ALIASES: dict[str, tuple[str, ...]] = {
    "medium": ("mediu", "medium", "tip mediu", "tip"),
    "publication_date": (
        "data",
        "data aparitiei",
        "data publicarii",
        "publication date",
        "date",
    ),
    "title": ("titlu", "title", "titlul articolului"),
    "source": ("sursa", "source", "publicatie", "publicatia"),
    "subject": ("subiect", "subject", "tema"),
    "audience": ("audienta", "audience"),
    "ave": ("ave", "valoare ave", "ave eur", "ad value equivalent"),
    "sentiment_original": ("tonalitate", "sentiment", "ton"),
    "importance_original": ("importanta", "importance", "relevanta"),
    "author": ("autor", "author", "jurnalist"),
    "county": ("judet", "county", "localitate"),
    "source_audience": ("audienta sursei", "source audience", "audienta publicatiei"),
    "subfolder_1": ("subfolder 1", "subfolder1", "sub folder 1", "folder 1", "categorie 1"),
    "subfolder_2": ("subfolder 2", "subfolder2", "sub folder 2", "folder 2", "categorie 2"),
    RETAILER_COLUMN_FIELD: ("retailer", "brand", "comerciant", "lant", "lant comercial", "client"),
}

_DIACRITICS_MAP = str.maketrans(
    {
        "ă": "a", "â": "a", "î": "i", "ș": "s", "ş": "s", "ț": "t", "ţ": "t",
        "Ă": "a", "Â": "a", "Î": "i", "Ș": "s", "Ş": "s", "Ț": "t", "Ţ": "t",
    }
)
_NON_ALNUM = re.compile(r"[^a-z0-9]+")


def normalize_header(text: Any) -> str:
    if text is None:
        return ""
    normalized = str(text).translate(_DIACRITICS_MAP).lower()
    return _NON_ALNUM.sub(" ", normalized).strip()


def _build_alias_lookup() -> dict[str, str]:
    lookup: dict[str, str] = {}
    for canonical_field, aliases in _RAW_ALIASES.items():
        for alias in aliases:
            lookup[normalize_header(alias)] = canonical_field
    return lookup


_ALIAS_TO_FIELD = _build_alias_lookup()

NUMERIC_FIELDS = {"audience", "ave", "source_audience"}
TEXT_FIELDS = {
    "medium",
    "subject",
    "sentiment_original",
    "importance_original",
    "author",
    "county",
    "subfolder_1",
    "subfolder_2",
}


class ParserError(Exception):
    pass


@dataclass
class ParsedRow:
    original_row_number: int
    retailer: str
    medium: str | None = None
    publication_date: date | None = None
    title: str | None = None
    article_url: str | None = None
    mediatrust_url: str | None = None
    source: str | None = None
    subject: str | None = None
    audience: float | None = None
    ave: float | None = None
    sentiment_original: str | None = None
    importance_original: str | None = None
    author: str | None = None
    county: str | None = None
    source_audience: float | None = None
    subfolder_1: str | None = None
    subfolder_2: str | None = None
    raw_json: dict = field(default_factory=dict)
    fingerprint: str = ""
    import_status: str = ImportStatus.VALID
    import_error: str | None = None
    retailer_confidence: str = ""
    retailer_review_status: str = RetailerReviewStatus.NEEDS_REVIEW
    retailer_raw_value: str | None = None


@dataclass
class ParseResult:
    workbook_sheet: str
    rows: list[ParsedRow]


def normalize_text_value(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def parse_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            excel_epoch = date(1899, 12, 30)
            return excel_epoch + timedelta(days=int(value))
        except (OverflowError, ValueError):
            return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%y", "%d/%m/%y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        return None
    return None


def parse_numeric(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        text = re.sub(r"[^0-9,.\-]", "", text)
        if not text or text in {"-", ".", ","}:
            return None
        if "," in text and "." in text:
            if text.rfind(",") > text.rfind("."):
                text = text.replace(".", "").replace(",", ".")
            else:
                text = text.replace(",", "")
        elif "," in text:
            text = text.replace(",", ".")
        try:
            return float(text)
        except ValueError:
            return None
    return None


def compute_fingerprint(
    retailer: str,
    title: str | None,
    source: str | None,
    publication_date: date | None,
    article_url: str | None,
) -> str:
    parts = [
        (retailer or "").strip().lower(),
        (title or "").strip().lower(),
        (source or "").strip().lower(),
        publication_date.isoformat() if publication_date else "",
        (article_url or "").strip().lower(),
    ]
    raw = "|".join(parts)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _find_header_row(worksheet: Worksheet) -> tuple[int, dict[int, str], dict[int, str]] | None:
    max_row = min(worksheet.max_row or 0, MAX_HEADER_SCAN_ROWS)
    for row in worksheet.iter_rows(min_row=1, max_row=max_row):
        header_map: dict[int, str] = {}
        header_labels: dict[int, str] = {}
        for cell in row:
            if cell.value is None:
                continue
            normalized = normalize_header(cell.value)
            canonical_field = _ALIAS_TO_FIELD.get(normalized)
            if canonical_field and cell.column not in header_map:
                header_map[cell.column] = canonical_field
                header_labels[cell.column] = str(cell.value).strip()
        distinct_fields = set(header_map.values()) - {RETAILER_COLUMN_FIELD}
        if len(distinct_fields) >= MIN_HEADER_MATCHES:
            return row[0].row, header_map, header_labels
    return None


def _select_worksheet(workbook) -> tuple[Worksheet, int, dict[int, str], dict[int, str]]:
    for worksheet in workbook.worksheets:
        if not worksheet.max_row or worksheet.max_row < 2:
            continue
        result = _find_header_row(worksheet)
        if result:
            header_row, header_map, header_labels = result
            return worksheet, header_row, header_map, header_labels
    raise ParserError("No worksheet with recognizable headers was found in this workbook.")


def _json_safe(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if value is None or isinstance(value, (int, float, str, bool)):
        return value
    return str(value)


def parse_workbook(
    file_path: str,
    original_filename: str,
    retailer_hint: str | None = None,
    retailer_hint_confirmed: bool = False,
) -> ParseResult:
    try:
        workbook = load_workbook(filename=file_path, data_only=True, read_only=False)
    except Exception as exc:  # noqa: BLE001 - surfaced as a parser error for the caller
        raise ParserError(f"Could not open workbook: {exc}") from exc

    worksheet, header_row, header_map, header_labels = _select_worksheet(workbook)

    # Tier-3 (file-level dominance) fallback needs to see every row's raw
    # mapped brand value up front — a lightweight pre-scan over just that
    # one column, before the main per-row loop applies the full decision
    # tree. Safe to iterate the worksheet twice: `read_only=False` above
    # loads the whole workbook into memory, unlike streaming mode.
    retailer_col_idx = next(
        (col for col, canonical_field in header_map.items() if canonical_field == RETAILER_COLUMN_FIELD),
        None,
    )
    dominant_value = None
    if retailer_col_idx is not None:
        raw_brand_values = [
            normalize_text_value(row[retailer_col_idx - 1].value)
            for row in worksheet.iter_rows(min_row=header_row + 1, max_row=worksheet.max_row)
        ]
        dominant_value = compute_dominant_retailer(raw_brand_values)

    rows: list[ParsedRow] = []
    for row in worksheet.iter_rows(min_row=header_row + 1, max_row=worksheet.max_row):
        row_number = row[0].row

        raw_json: dict[str, Any] = {}
        has_any_value = False
        for cell in row:
            if cell.value is not None and str(cell.value).strip() != "":
                has_any_value = True
            label = header_labels.get(cell.column, f"col_{cell.column}")
            raw_json[label] = _json_safe(cell.value)

        if not has_any_value:
            continue

        field_cells: dict[str, Any] = {}
        for col_idx, canonical_field in header_map.items():
            field_cells[canonical_field] = row[col_idx - 1]

        title_cell = field_cells.get("title")
        source_cell = field_cells.get("source")

        title = normalize_text_value(title_cell.value) if title_cell is not None else None
        mediatrust_url = (
            title_cell.hyperlink.target
            if title_cell is not None and title_cell.hyperlink
            else None
        )
        source = normalize_text_value(source_cell.value) if source_cell is not None else None
        article_url = (
            source_cell.hyperlink.target
            if source_cell is not None and source_cell.hyperlink
            else None
        )

        mapped_brand_value = None
        retailer_column_cell = field_cells.get(RETAILER_COLUMN_FIELD)
        if retailer_column_cell is not None:
            mapped_brand_value = normalize_text_value(retailer_column_cell.value)

        assignment = assign_retailer(
            mapped_brand_value,
            retailer_hint=retailer_hint,
            retailer_hint_confirmed=retailer_hint_confirmed,
            filename=original_filename,
            dominant_value=dominant_value,
        )
        retailer = assignment.value

        publication_date = None
        if "publication_date" in field_cells:
            publication_date = parse_date(field_cells["publication_date"].value)

        numeric_values: dict[str, float | None] = {}
        for numeric_field in NUMERIC_FIELDS:
            cell = field_cells.get(numeric_field)
            numeric_values[numeric_field] = parse_numeric(cell.value) if cell is not None else None

        text_values: dict[str, str | None] = {}
        for text_field in TEXT_FIELDS:
            cell = field_cells.get(text_field)
            text_values[text_field] = (
                normalize_text_value(cell.value) if cell is not None else None
            )

        if title or source:
            import_status = ImportStatus.VALID
            import_error = None
        else:
            import_status = ImportStatus.INVALID
            import_error = "Missing both title and source."

        fingerprint = compute_fingerprint(retailer, title, source, publication_date, article_url)

        rows.append(
            ParsedRow(
                original_row_number=row_number,
                retailer=retailer,
                medium=text_values.get("medium"),
                publication_date=publication_date,
                title=title,
                article_url=article_url,
                mediatrust_url=mediatrust_url,
                source=source,
                subject=text_values.get("subject"),
                audience=numeric_values.get("audience"),
                ave=numeric_values.get("ave"),
                sentiment_original=text_values.get("sentiment_original"),
                importance_original=text_values.get("importance_original"),
                author=text_values.get("author"),
                county=text_values.get("county"),
                source_audience=numeric_values.get("source_audience"),
                subfolder_1=text_values.get("subfolder_1"),
                subfolder_2=text_values.get("subfolder_2"),
                raw_json=raw_json,
                fingerprint=fingerprint,
                import_status=import_status,
                import_error=import_error,
                retailer_confidence=assignment.confidence,
                retailer_review_status=(
                    RetailerReviewStatus.NEEDS_REVIEW
                    if assignment.needs_review
                    else RetailerReviewStatus.CONFIRMED
                ),
                retailer_raw_value=assignment.raw_value,
            )
        )

    return ParseResult(workbook_sheet=worksheet.title, rows=rows)
