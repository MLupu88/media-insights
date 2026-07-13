"""MSL brand styling constants and cell/sheet helpers for XLSX report
generation (Phase 6C). Same brand tokens as `report_pptx_style.py`
(`tailwind.config.js`) applied through openpyxl instead of python-pptx.
"""

from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.services.report_pptx_style import LOGO_PATH

INK = "111111"
PAPER = "FAF9F6"
LINE = "E4E1DA"
ACCENT = "B3492B"
WHITE = "FFFFFF"

SANS_FONT = "Inter"

MAX_SHEET_NAME_CHARS = 31  # hard Excel limit

# Logo is 420x60; scaled down to a small in-sheet header mark, never large
# enough to spill into the data columns (A/B) that every metadata block uses.
_LOGO_HEIGHT_PX = 32
_LOGO_WIDTH_PX = int(_LOGO_HEIGHT_PX * 420 / 60)
BRAND_HEADER_ROWS = 2  # rows reserved by add_branded_header before content may start


def safe_sheet_name(name: str) -> str:
    """Excel silently refuses/corrupts a workbook with a sheet name over 31
    characters — truncate defensively so a future long sheet title can
    never reintroduce that failure mode.
    """
    return name[:MAX_SHEET_NAME_CHARS]

HEADER_FONT = Font(name=SANS_FONT, bold=True, color=WHITE)
HEADER_FILL = PatternFill(start_color=INK, end_color=INK, fill_type="solid")
LABEL_FONT = Font(name=SANS_FONT, bold=True)
BODY_FONT = Font(name=SANS_FONT)
TITLE_FONT = Font(name=SANS_FONT, bold=True, size=14, color=ACCENT)

BRAND_BAND_FILL = PatternFill(start_color=INK, end_color=INK, fill_type="solid")
BRAND_BAND_FONT = Font(name=SANS_FONT, bold=True, size=13, color=WHITE)


def add_branded_header(ws: Worksheet, title: str, band_cols: int = 3) -> int:
    """Writes an ink-filled title band across `band_cols` columns of row 1
    with the MSL wordmark anchored just to its right, then returns the
    first free row below the band for subsequent content.

    The logo is anchored (not embedded in a cell), so it never obscures a
    cell value and never affects `freeze_panes`/`auto_filter`, both of
    which are computed from row/column indices alone. It's placed clear of
    every sheet's data columns (A/B here, occasionally C), so it can never
    overlap real content either.
    """
    ws.row_dimensions[1].height = 24
    for col in range(1, band_cols + 1):
        ws.cell(row=1, column=col).fill = BRAND_BAND_FILL
    ws.cell(row=1, column=1, value=title).font = BRAND_BAND_FONT

    image = XLImage(LOGO_PATH)
    image.width = _LOGO_WIDTH_PX
    image.height = _LOGO_HEIGHT_PX
    ws.add_image(image, f"{get_column_letter(band_cols + 1)}1")

    return BRAND_HEADER_ROWS + 1


def write_title(ws: Worksheet, text: str, row: int = 1) -> int:
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = TITLE_FONT
    return row + 2


def write_metadata_block(ws: Worksheet, pairs: list[tuple[str, str]], start_row: int = 1) -> int:
    row = start_row
    for label, value in pairs:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=2, value=value).font = BODY_FONT
        row += 1
    return row + 1  # one blank row after the block


def write_header_row(ws: Worksheet, row: int, headers: list[str]) -> None:
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_index, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")


def freeze_header(ws: Worksheet, header_row: int = 1) -> None:
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate


def apply_autofilter(ws: Worksheet, header_row: int, n_cols: int, n_data_rows: int) -> None:
    last_col = get_column_letter(n_cols)
    last_row = header_row + max(n_data_rows, 1)
    ws.auto_filter.ref = f"A{header_row}:{last_col}{last_row}"


def set_column_widths(ws: Worksheet, widths: list[int]) -> None:
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def set_percent(cell, value: float | None) -> None:
    """`value` is a percentage number already computed by the backend
    (e.g. `24.6` meaning 24.6%) — converted to a true Excel fraction so
    the native `"0.0%"` format displays correctly. Writing `24.6` directly
    under a `"0.0%"` format would render as `"2460.0%"`, silently
    corrupting the figure — this is the one place that mistake must never
    happen.
    """
    if value is None:
        cell.value = None
        return
    cell.value = value / 100
    cell.number_format = "0.0%"


def set_pp(cell, value: float | None) -> None:
    """Percentage-point values are never given Excel's native `%` format
    — always a plain 1-decimal number, under a header that says "(pp)"
    explicitly, so a percent and a percentage-point column can never be
    visually confused.
    """
    cell.value = value
    cell.number_format = "0.0"


def set_number(cell, value, decimals: int = 0) -> None:
    cell.value = value
    cell.number_format = "#,##0" if decimals == 0 else f"#,##0.{'0' * decimals}"
