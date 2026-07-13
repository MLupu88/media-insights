"""Brand-detection decision tree (approved reporting-scope plan, §8, with
the safety correction from the follow-up review). Most of this file tests
`assign_retailer`/`compute_dominant_retailer` directly — fast, precise unit
tests with full control over every input — plus a handful of integration
tests through `parse_workbook` and the real upload route to prove the
wiring, multi-brand-per-file support, and dedup behavior end to end.
"""

from openpyxl import Workbook

from app.models.article import Article, RetailerConfidence, RetailerReviewStatus
from app.models.project import Project
from app.services.excel_parser import parse_workbook
from app.services.retailers import (
    DOMINANCE_THRESHOLD_PCT,
    MIN_DOMINANCE_SAMPLE_SIZE,
    assign_retailer,
    compute_dominant_retailer,
)

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# --- assign_retailer: non-empty row value ------------------------------------


def test_explicit_row_brand_wins_even_with_a_conflicting_filename():
    result = assign_retailer("Auchan", filename="Carrefour_Q2.xlsx")
    assert result.value == "Auchan"
    assert result.confidence == RetailerConfidence.EXPLICIT_COLUMN
    assert result.needs_review is False


def test_explicit_row_brand_wins_even_with_a_confirmed_hint_for_another_brand():
    result = assign_retailer("Auchan", retailer_hint="Carrefour", retailer_hint_confirmed=True)
    assert result.value == "Auchan"
    assert result.confidence == RetailerConfidence.EXPLICIT_COLUMN


def test_explicit_row_brand_wins_even_with_file_level_dominance_for_another_brand():
    result = assign_retailer("Auchan", dominant_value="Carrefour")
    assert result.value == "Auchan"
    assert result.confidence == RetailerConfidence.EXPLICIT_COLUMN


def test_non_empty_unrecognized_row_brand_goes_to_review_and_is_never_overridden():
    """The central safety rule: a present-but-unmatched value is never
    guessed past, even when every weaker signal points somewhere else.
    """
    result = assign_retailer(
        "Regional Chain XYZ",
        retailer_hint="Carrefour",
        retailer_hint_confirmed=True,
        filename="Auchan_Q2.xlsx",
        dominant_value="Auchan",
    )
    assert result.value == "unknown"
    assert result.confidence == RetailerConfidence.NEEDS_REVIEW
    assert result.needs_review is True
    assert result.raw_value == "Regional Chain XYZ"


def test_filename_fallback_never_applies_to_a_non_blank_row_value():
    result = assign_retailer("Not A Retailer", filename="Auchan_Q2.xlsx")
    assert result.value != "Auchan"
    assert result.confidence == RetailerConfidence.NEEDS_REVIEW


# --- assign_retailer: blank row value ----------------------------------------


def test_blank_row_uses_confirmed_hint_first():
    result = assign_retailer(
        None,
        retailer_hint="Carrefour",
        retailer_hint_confirmed=True,
        filename="Auchan.xlsx",
        dominant_value="Lidl",
    )
    assert result.value == "Carrefour"
    assert result.confidence == RetailerConfidence.CONFIRMED_MAPPING


def test_blank_row_ignores_an_unconfirmed_hint():
    result = assign_retailer(
        None, retailer_hint="Carrefour", retailer_hint_confirmed=False, filename="Auchan.xlsx"
    )
    assert result.value == "Auchan"
    assert result.confidence == RetailerConfidence.FILENAME_FALLBACK


def test_blank_row_uses_file_level_dominance_when_no_confirmed_hint():
    result = assign_retailer(None, dominant_value="Lidl", filename="Auchan.xlsx")
    assert result.value == "Lidl"
    assert result.confidence == RetailerConfidence.FILE_LEVEL_INFERENCE


def test_blank_row_falls_back_to_filename_when_no_dominance():
    result = assign_retailer(None, dominant_value=None, filename="Auchan_Q2.xlsx")
    assert result.value == "Auchan"
    assert result.confidence == RetailerConfidence.FILENAME_FALLBACK


def test_blank_row_with_no_signal_at_all_goes_to_review():
    result = assign_retailer(None)
    assert result.value == "unknown"
    assert result.confidence == RetailerConfidence.NEEDS_REVIEW
    assert result.needs_review is True
    assert result.raw_value is None


# --- compute_dominant_retailer: named-constant thresholds --------------------


def test_dominance_requires_the_minimum_sample_size():
    values = ["Auchan"] * (MIN_DOMINANCE_SAMPLE_SIZE - 1)
    assert compute_dominant_retailer(values) is None


def test_dominance_applies_at_exactly_the_minimum_sample_size_and_full_agreement():
    values = ["Auchan"] * MIN_DOMINANCE_SAMPLE_SIZE
    assert compute_dominant_retailer(values) == "Auchan"


def test_dominance_requires_the_agreement_threshold():
    values = ["Auchan"] * 5 + ["Carrefour"] * 5  # 50% agreement
    assert compute_dominant_retailer(values) is None


def test_dominance_applies_at_exactly_the_agreement_threshold():
    assert DOMINANCE_THRESHOLD_PCT == 90
    values = ["Auchan"] * 9 + ["Carrefour"] * 1  # exactly 90% agreement
    assert compute_dominant_retailer(values) == "Auchan"


def test_dominance_just_below_the_agreement_threshold_is_rejected():
    values = ["Auchan"] * 8 + ["Carrefour"] * 2  # 80% agreement
    assert compute_dominant_retailer(values) is None


def test_dominance_ignores_blank_and_unrecognized_values_in_the_sample():
    values = ["Auchan"] * 5 + [None, "", "Not A Retailer"]
    assert compute_dominant_retailer(values) == "Auchan"


# --- parse_workbook integration: multi-brand files ---------------------------


def _build_brand_column_workbook(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Monitorizare"
    headers = ["Titlu", "Sursa", "Data", "Subiect", "Brand"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    for offset, row in enumerate(rows, start=1):
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1 + offset, column=col, value=row.get(header))
    wb.save(path)
    return path


def test_one_file_can_contain_several_brands(tmp_path):
    path = _build_brand_column_workbook(
        tmp_path / "Multi Brand.xlsx",
        [
            {"Titlu": "Auchan story", "Sursa": "Ziarul", "Data": "01.04.2026", "Subiect": "S", "Brand": "Auchan"},
            {"Titlu": "Carrefour story", "Sursa": "Ziarul", "Data": "02.04.2026", "Subiect": "S", "Brand": "Carrefour"},
            {"Titlu": "Penny story", "Sursa": "Ziarul", "Data": "03.04.2026", "Subiect": "S", "Brand": "Penny"},
        ],
    )

    result = parse_workbook(str(path), "Mixed Retailers Q2 2026.xlsx")

    retailers = {row.retailer for row in result.rows}
    assert retailers == {"Auchan", "Carrefour", "Penny / Rewe"}
    assert all(row.retailer_confidence == "explicit_column" for row in result.rows)
    assert all(row.retailer_review_status == "confirmed" for row in result.rows)


def test_non_empty_unrecognized_brand_in_a_real_workbook_goes_to_review(tmp_path):
    path = _build_brand_column_workbook(
        tmp_path / "Unknown Brand.xlsx",
        [{"Titlu": "Story", "Sursa": "Ziarul", "Data": "01.04.2026", "Subiect": "S", "Brand": "Some Local Shop"}],
    )

    result = parse_workbook(str(path), "Auchan Q2 2026.xlsx")  # filename WOULD match Auchan

    row = result.rows[0]
    assert row.retailer == "unknown"
    assert row.retailer_confidence == "needs_review"
    assert row.retailer_review_status == "needs_review"
    assert row.retailer_raw_value == "Some Local Shop"


# --- full pipeline: dedup across brands within one multi-brand file ----------


def test_dedup_works_correctly_across_different_brands_in_one_file(
    authenticated_client, db_session, tmp_path
):
    response = authenticated_client.post(
        "/projects", data={"name": "Multi Brand Dedup", "quarter": "2026-Q2"}, follow_redirects=False
    )
    assert response.status_code == 303
    project = db_session.query(Project).filter_by(name="Multi Brand Dedup").one()

    # Same title/source/date, but two different resolved brands — the
    # fingerprint includes retailer, so these must NOT dedupe against each
    # other; a third row exactly matching the first (same brand too) must.
    path = _build_brand_column_workbook(
        tmp_path / "Dedup.xlsx",
        [
            {"Titlu": "Shared headline", "Sursa": "Ziarul", "Data": "01.04.2026", "Subiect": "S", "Brand": "Auchan"},
            {"Titlu": "Shared headline", "Sursa": "Ziarul", "Data": "01.04.2026", "Subiect": "S", "Brand": "Carrefour"},
            {"Titlu": "Shared headline", "Sursa": "Ziarul", "Data": "01.04.2026", "Subiect": "S", "Brand": "Auchan"},
        ],
    )

    response = authenticated_client.post(
        f"/projects/{project.id}/files",
        files=[("files", ("Multi Brand.xlsx", path.read_bytes(), XLSX_CONTENT_TYPE))],
    )
    assert response.status_code == 200

    articles = (
        db_session.query(Article)
        .filter_by(project_id=project.id)
        .order_by(Article.original_row_number)
        .all()
    )
    assert len(articles) == 3

    auchan_rows = [a for a in articles if a.retailer == "Auchan"]
    carrefour_rows = [a for a in articles if a.retailer == "Carrefour"]
    assert len(auchan_rows) == 2
    assert len(carrefour_rows) == 1

    assert sum(1 for a in auchan_rows if a.is_duplicate) == 1
    assert carrefour_rows[0].is_duplicate is False


# --- legacy/backfilled rows ---------------------------------------------------


def test_articles_created_without_explicit_provenance_default_to_legacy_confirmed(
    project_factory, article_factory
):
    """Mirrors what the migration's backfill does for pre-existing rows —
    any `Article` created without explicitly setting the new provenance
    fields lands on the same 'legacy'/'confirmed' defaults, never in the
    Review population.
    """
    project = project_factory()
    articles = article_factory(project, count=1, retailer="Auchan")
    article = articles[0]

    assert article.retailer_confidence == RetailerConfidence.LEGACY
    assert article.retailer_review_status == RetailerReviewStatus.CONFIRMED
    assert article.retailer_raw_value is None
