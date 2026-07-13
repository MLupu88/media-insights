"""Regression tests for the production Excel-import hotfix:

- the `author` (and sibling) field-length production incident is fixed at
  the schema level (Text, not a bounded VARCHAR) -- proven with a real
  >255-character value, no mocking needed;
- a genuine database failure during article insertion (flush/commit) is
  caught, rolled back, and produces a safe, readable result -- never a raw
  500 -- proven by mocking `Session.commit` to fail exactly on the
  article-insertion commit, since the schema fix above means a real
  >255-char value can no longer reproduce a database-level failure;
- the outer per-request exception handler in `app/api/files.py` captures
  `batch_id`/`uploaded_file_id` as plain UUIDs before any risky operation,
  so `PendingRollbackError` (the reported secondary error) cannot occur;
- batch finalization (failed vs partially_completed) is correct when one
  file succeeds and another fails;
- no raw SQL, parameter payloads, or workbook content ever reach a
  user-visible error message.
"""

from datetime import date
from io import BytesIO
from unittest.mock import patch

from openpyxl import Workbook
from sqlalchemy.exc import DataError
from sqlalchemy.orm import Session

from app.models.article import Article
from app.models.import_batch import ImportBatch, ImportBatchStatus
from app.models.project import Project
from app.models.uploaded_file import UploadedFile, UploadedFileStatus

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _create_project(client, db_session, name="Import Error Test", quarter="2026-Q2") -> Project:
    response = client.post(
        "/projects", data={"name": name, "quarter": quarter}, follow_redirects=False
    )
    assert response.status_code == 303
    return db_session.query(Project).filter_by(name=name).one()


def _build_workbook_with_long_author(author_length: int) -> bytes:
    """One valid row, with `Autor` deliberately longer than the old 255-char
    VARCHAR limit -- reproduces the exact shape of the production incident
    (Carrefour Q2 2026, row 72, a 405-character author value) without
    depending on any real customer workbook.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Monitorizare"
    headers = ["Titlu", "Sursa", "Mediu", "Autor", "Data"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    ws.cell(row=2, column=1, value="A long-author regression test article")
    ws.cell(row=2, column=2, value="Regression Test Source")
    ws.cell(row=2, column=3, value="Online")
    ws.cell(row=2, column=4, value="A" * author_length)
    ws.cell(row=2, column=5, value=date(2026, 5, 1))
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# --- Part 2 proof: the schema fix itself -------------------------------------


def test_source_value_longer_than_255_characters_is_stored_without_error(
    authenticated_client, db_session
):
    project = _create_project(authenticated_client, db_session)
    long_author = "A" * 405  # matches the real production incident's length

    response = authenticated_client.post(
        f"/projects/{project.id}/files",
        files=[("files", ("Long Author Test.xlsx", _build_workbook_with_long_author(405), XLSX_CONTENT_TYPE))],
    )

    assert response.status_code == 200
    uploaded_file = db_session.query(UploadedFile).filter_by(project_id=project.id).one()
    assert uploaded_file.status == UploadedFileStatus.COMPLETED
    assert uploaded_file.error_message is None

    article = db_session.query(Article).filter_by(uploaded_file_id=uploaded_file.id).one()
    assert article.author == long_author
    assert len(article.author) == 405


# --- database failure during article insertion -------------------------------


def _patched_commit_failing_on_article_insert(original_commit):
    """Raises a `DataError` (matching the real
    `StringDataRightTruncation`/`character varying(255)` production error)
    on exactly the commit that has a pending `Article` insert -- robust
    against how many other commits happen before it, unlike a fixed call
    count.
    """

    def _commit(self, *args, **kwargs):
        if any(isinstance(obj, Article) for obj in self.new):
            raise DataError(
                "INSERT INTO articles (author, ...) VALUES (%(author)s, ...)",
                {"author": "some very long value that would never be shown to a user"},
                Exception("value too long for type character varying(255)"),
            )
        return original_commit(self, *args, **kwargs)

    return _commit


def test_database_failure_during_article_insertion_is_handled_gracefully(
    authenticated_client, db_session, standard_workbook_path
):
    project = _create_project(authenticated_client, db_session)
    original_commit = Session.commit

    with patch.object(Session, "commit", _patched_commit_failing_on_article_insert(original_commit)):
        response = authenticated_client.post(
            f"/projects/{project.id}/files",
            files=[("files", ("Auchan Q2 2026.xlsx", standard_workbook_path.read_bytes(), XLSX_CONTENT_TYPE))],
        )

    # Never a raw 500 for a row-level database error.
    assert response.status_code != 500
    assert "Traceback" not in response.text
    assert "internal server error" not in response.text.lower()

    uploaded_file = db_session.query(UploadedFile).filter_by(project_id=project.id).one()
    assert uploaded_file.status == UploadedFileStatus.FAILED
    assert uploaded_file.row_count == 0
    assert uploaded_file.valid_row_count == 0
    assert uploaded_file.invalid_row_count == 0
    assert uploaded_file.duplicate_row_count == 0
    assert uploaded_file.error_message
    assert "exceed the database" in uploaded_file.error_message

    # No Article rows were left behind from the rolled-back insert.
    assert db_session.query(Article).filter_by(uploaded_file_id=uploaded_file.id).count() == 0


def test_database_failure_recomputes_project_totals_from_successful_files_only(
    authenticated_client, db_session, standard_workbook_path, penny_workbook_path
):
    project = _create_project(authenticated_client, db_session)
    original_commit = Session.commit

    # First file succeeds normally.
    authenticated_client.post(
        f"/projects/{project.id}/files",
        files=[("files", ("Auchan Q2 2026.xlsx", standard_workbook_path.read_bytes(), XLSX_CONTENT_TYPE))],
    )
    db_session.refresh(project)
    rows_after_first_file = project.total_rows
    assert rows_after_first_file > 0

    # Second file fails during article insertion.
    with patch.object(Session, "commit", _patched_commit_failing_on_article_insert(original_commit)):
        authenticated_client.post(
            f"/projects/{project.id}/files",
            files=[("files", ("Penny - Rewe Q2 2026.xlsx", penny_workbook_path.read_bytes(), XLSX_CONTENT_TYPE))],
        )

    db_session.refresh(project)
    # Totals reflect only the successful file -- the failed file's rolled-back
    # rows never contribute.
    assert project.total_rows == rows_after_first_file


# --- outer handler: batch_id captured before any risky operation -------------


def test_rollback_happens_before_batch_is_reloaded_on_unexpected_failure(
    authenticated_client, db_session, standard_workbook_path
):
    """Forces an exception to escape `import_uploaded_file` entirely
    (simulating a bug outside the row-level DB-error path already handled
    internally) -- proves the outer handler's `fail_import_batch(db,
    batch_id, ...)` call (a plain UUID, captured before the try block)
    completes without a secondary `PendingRollbackError`, and the request
    still returns a readable page, not a raw 500.
    """
    project = _create_project(authenticated_client, db_session)

    with patch(
        "app.api.files.import_uploaded_file", side_effect=RuntimeError("unexpected bug")
    ):
        response = authenticated_client.post(
            f"/projects/{project.id}/files",
            files=[("files", ("Auchan Q2 2026.xlsx", standard_workbook_path.read_bytes(), XLSX_CONTENT_TYPE))],
        )

    assert response.status_code == 500
    assert "Traceback" not in response.text
    assert "PendingRollbackError" not in response.text
    assert "unexpected bug" not in response.text

    batch = db_session.query(ImportBatch).filter_by(project_id=project.id).one()
    assert batch.status == ImportBatchStatus.FAILED
    assert batch.completed_at is not None


# --- one success + one failure -> partially_completed -------------------------


def test_one_successful_file_followed_by_one_failed_file_is_partially_completed(
    authenticated_client, db_session, standard_workbook_path, penny_workbook_path
):
    project = _create_project(authenticated_client, db_session)
    original_commit = Session.commit
    call_state = {"first_article_commit_done": False}

    def commit_fails_on_second_files_articles(self, *args, **kwargs):
        has_pending_articles = any(isinstance(obj, Article) for obj in self.new)
        if has_pending_articles and call_state["first_article_commit_done"]:
            raise DataError(
                "INSERT INTO articles ...", {}, Exception("value too long for type character varying(255)")
            )
        if has_pending_articles:
            call_state["first_article_commit_done"] = True
        return original_commit(self, *args, **kwargs)

    with patch.object(Session, "commit", commit_fails_on_second_files_articles):
        response = authenticated_client.post(
            f"/projects/{project.id}/files",
            files=[
                ("files", ("Auchan Q2 2026.xlsx", standard_workbook_path.read_bytes(), XLSX_CONTENT_TYPE)),
                ("files", ("Penny - Rewe Q2 2026.xlsx", penny_workbook_path.read_bytes(), XLSX_CONTENT_TYPE)),
            ],
        )

    assert response.status_code == 200

    uploaded_files = db_session.query(UploadedFile).filter_by(project_id=project.id).all()
    statuses = {f.original_filename: f.status for f in uploaded_files}
    assert statuses["Auchan Q2 2026.xlsx"] == UploadedFileStatus.COMPLETED
    assert statuses["Penny - Rewe Q2 2026.xlsx"] == UploadedFileStatus.FAILED

    batch = db_session.query(ImportBatch).filter_by(project_id=project.id).one()
    assert batch.status == ImportBatchStatus.PARTIALLY_COMPLETED
    assert batch.files_accepted == 1
    assert batch.files_rejected == 1
    assert batch.completed_at is not None
    assert batch.error_summary
    assert "INSERT INTO" not in batch.error_summary


# --- no raw SQL or workbook content in user-visible errors --------------------


def test_user_visible_error_never_contains_raw_sql_or_workbook_content(
    authenticated_client, db_session, standard_workbook_path
):
    project = _create_project(authenticated_client, db_session)
    original_commit = Session.commit

    with patch.object(Session, "commit", _patched_commit_failing_on_article_insert(original_commit)):
        response = authenticated_client.post(
            f"/projects/{project.id}/files",
            files=[("files", ("Auchan Q2 2026.xlsx", standard_workbook_path.read_bytes(), XLSX_CONTENT_TYPE))],
        )

    forbidden_fragments = [
        "INSERT INTO",
        "VALUES (",
        "some very long value that would never be shown to a user",
        "%(author)s",
    ]
    for fragment in forbidden_fragments:
        assert fragment not in response.text

    uploaded_file = db_session.query(UploadedFile).filter_by(project_id=project.id).one()
    for fragment in forbidden_fragments:
        assert fragment not in (uploaded_file.error_message or "")


# --- retry works after the schema migration -----------------------------------


def test_retry_succeeds_for_a_file_with_a_long_author_value(authenticated_client, db_session):
    """Before the schema fix, a file with a >255-char author value would
    fail every retry identically (same schema, same error). After the fix,
    a first attempt that failed for an unrelated reason (simulated here via
    a corrupt upload) can be retried and now succeeds even though the
    underlying content includes a value that used to be untouchable.
    """
    project = _create_project(authenticated_client, db_session)

    broken_bytes = b"this is not a real xlsx file"
    authenticated_client.post(
        f"/projects/{project.id}/files",
        files=[("files", ("Long Author Retry.xlsx", broken_bytes, XLSX_CONTENT_TYPE))],
    )
    uploaded_file = db_session.query(UploadedFile).filter_by(project_id=project.id).one()
    assert uploaded_file.status == UploadedFileStatus.FAILED

    with open(uploaded_file.stored_path, "wb") as f:
        f.write(_build_workbook_with_long_author(405))

    response = authenticated_client.post(
        f"/projects/{project.id}/files/{uploaded_file.id}/retry", follow_redirects=False
    )
    assert response.status_code == 303

    db_session.refresh(uploaded_file)
    assert uploaded_file.status == UploadedFileStatus.COMPLETED
    article = db_session.query(Article).filter_by(uploaded_file_id=uploaded_file.id).one()
    assert len(article.author) == 405
