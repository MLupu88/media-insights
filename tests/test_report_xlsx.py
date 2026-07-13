from io import BytesIO
from unittest.mock import patch

import openpyxl
import pytest

from app.services import report_xlsx
from app.services.analytics import AnalyticsFilters
from app.services.report_contract import EXCEL_TOP_N
from app.services.report_data import (
    ReportTooLargeError,
    build_comparison_report_data,
    build_project_report_data,
)


def _open(data: bytes):
    return openpyxl.load_workbook(BytesIO(data))


def _all_cells(wb):
    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            yield from row


def test_project_workbook_sheet_names(report_db, project_factory, article_factory):
    project = project_factory()
    article_factory(project, count=1, retailer="Auchan")

    data = build_project_report_data(report_db, project.id, AnalyticsFilters())
    wb = _open(report_xlsx.build_project_xlsx(data))

    assert wb.sheetnames == [
        "Executive Summary", "KPI Summary", "Brand Performance", "Topic Distribution",
        "Sentiment", "Publications", "Story Clusters", "Validated Insights",
        "Article Detail", "Methodology",
    ]


def test_comparison_workbook_sheet_names(report_db, project_factory, article_factory):
    a = project_factory(name="A", quarter="2026-Q1")
    b = project_factory(name="B", quarter="2026-Q2")
    article_factory(a, count=1, retailer="Auchan")
    article_factory(b, count=1, retailer="Auchan")

    data = build_comparison_report_data(report_db, [a.id], [b.id], AnalyticsFilters())
    wb = _open(report_xlsx.build_comparison_xlsx(data))

    assert wb.sheetnames == [
        "Executive Summary", "KPI Comparison", "Brand Performance Comparison",
        "Topic & Category Shifts", "Sentiment & Brand Role", "Publication Movement",
        "Story Movement", "Concentration & Volatility", "Validated Insights",
        "Article Detail", "Methodology",
    ]
    assert all(len(name) <= 31 for name in wb.sheetnames)


def test_no_formulas_anywhere(report_db, db_session, project_factory, article_factory, classification_factory):
    project = project_factory()
    articles = article_factory(project, count=3, retailer="Auchan")
    classification_factory(articles[0])

    data = build_project_report_data(report_db, project.id, AnalyticsFilters())
    wb = _open(report_xlsx.build_project_xlsx(data))

    formula_cells = [cell.coordinate for cell in _all_cells(wb) if cell.data_type == "f"]
    assert formula_cells == []


def test_no_formulas_anywhere_comparison(report_db, project_factory, article_factory):
    a = project_factory(name="A", quarter="2026-Q1")
    b = project_factory(name="B", quarter="2026-Q2")
    article_factory(a, count=1, retailer="Auchan")
    article_factory(b, count=1, retailer="Auchan")

    data = build_comparison_report_data(report_db, [a.id], [b.id], AnalyticsFilters())
    wb = _open(report_xlsx.build_comparison_xlsx(data))

    formula_cells = [cell.coordinate for cell in _all_cells(wb) if cell.data_type == "f"]
    assert formula_cells == []


def test_brand_sheet_frozen_and_filtered(report_db, project_factory, article_factory):
    project = project_factory()
    article_factory(project, count=1, retailer="Auchan")

    data = build_project_report_data(report_db, project.id, AnalyticsFilters())
    wb = _open(report_xlsx.build_project_xlsx(data))

    ws = wb["Brand Performance"]
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref is not None


def test_percent_columns_use_native_percent_format(report_db, project_factory, article_factory):
    project = project_factory()
    article_factory(project, count=1, retailer="Auchan")

    data = build_project_report_data(report_db, project.id, AnalyticsFilters())
    wb = _open(report_xlsx.build_project_xlsx(data))

    ws = wb["Brand Performance"]
    sov_cell = ws.cell(row=2, column=3)
    assert sov_cell.number_format == "0.0%"
    # value is the true fraction, not the raw percentage number.
    assert 0 <= sov_cell.value <= 1


def test_pp_columns_never_use_percent_format(report_db, project_factory, article_factory):
    a = project_factory(name="A", quarter="2026-Q1")
    b = project_factory(name="B", quarter="2026-Q2")
    article_factory(a, count=1, retailer="Auchan")
    article_factory(b, count=1, retailer="Auchan")

    data = build_comparison_report_data(report_db, [a.id], [b.id], AnalyticsFilters())
    wb = _open(report_xlsx.build_comparison_xlsx(data))

    ws = wb["Brand Performance Comparison"]
    headers = [c.value for c in ws[1]]
    pp_col = headers.index("SOV Delta (pp)") + 1
    pp_cell = ws.cell(row=2, column=pp_col)
    assert "%" not in pp_cell.number_format
    assert pp_cell.number_format == "0.0"


def test_article_detail_includes_urls_and_is_never_truncated_by_label(
    report_db, project_factory, article_factory
):
    long_title = "B" * 60
    project = project_factory()
    article_factory(
        project, count=1, retailer="Auchan", title=long_title,
        article_url="https://example.test/a", mediatrust_url="https://mediatrust.test/a",
    )

    data = build_project_report_data(report_db, project.id, AnalyticsFilters())
    wb = _open(report_xlsx.build_project_xlsx(data))

    ws = wb["Article Detail"]
    headers = [c.value for c in ws[1]]
    assert "Article URL" in headers
    assert "Mediatrust URL" in headers
    title_col = headers.index("Title") + 1
    assert ws.cell(row=2, column=title_col).value == long_title  # never truncated


def test_comparison_article_detail_has_period_column(report_db, project_factory, article_factory):
    a = project_factory(name="A", quarter="2026-Q1")
    b = project_factory(name="B", quarter="2026-Q2")
    article_factory(a, count=1, retailer="Auchan")
    article_factory(b, count=1, retailer="Auchan")

    data = build_comparison_report_data(report_db, [a.id], [b.id], AnalyticsFilters())
    wb = _open(report_xlsx.build_comparison_xlsx(data))

    ws = wb["Article Detail"]
    headers = [c.value for c in ws[1]]
    assert headers[0] == "Period"
    periods = {ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)}
    assert periods == {"Baseline", "Comparison"}


def test_validated_insights_sheet_excludes_rejected(
    report_db, db_session, project_factory, article_factory
):
    from app.models.narrative import NarrativeInsight, NarrativeValidationStatus
    from app.services.narrative_service import create_project_generation

    project = project_factory()
    article_factory(project, count=1, retailer="Auchan")
    generation, _ = create_project_generation(db_session, project, AnalyticsFilters())
    generation.status = "complete"
    db_session.commit()

    valid = NarrativeInsight(
        generation_id=generation.id, project_id=generation.project_id, narrative_type="recommendations",
        key="r1", title="Valid Rec", narrative="Do this.", evidence_type="kpi_delta", evidence=[],
        raw_candidate={}, validation_status=NarrativeValidationStatus.VALID,
    )
    rejected = NarrativeInsight(
        generation_id=generation.id, project_id=generation.project_id, narrative_type="key_findings",
        key="k1", title="Rejected Finding", narrative="Bad text.", evidence_type="kpi_delta", evidence=[],
        raw_candidate={}, validation_status=NarrativeValidationStatus.REJECTED, rejection_reason="bad",
    )
    db_session.add_all([valid, rejected])
    db_session.commit()

    data = build_project_report_data(report_db, project.id, AnalyticsFilters())
    wb = _open(report_xlsx.build_project_xlsx(data))

    ws = wb["Validated Insights"]
    titles = [ws.cell(row=r, column=3).value for r in range(2, ws.max_row + 1)]
    assert "Valid Rec" in titles
    assert "Rejected Finding" not in titles
    type_col = [c.value for c in ws[1]].index("Type") + 1
    assert ws.cell(row=2, column=type_col).value == "Recommendation"


def test_methodology_sheet_coverage_matches_data_layer(report_db, project_factory, article_factory):
    project = project_factory()
    article_factory(project, count=3, retailer="Auchan")

    with patch("app.services.report_data.MAX_ARTICLE_DETAIL_ROWS", 1):
        data = build_project_report_data(report_db, project.id, AnalyticsFilters())
    wb = _open(report_xlsx.build_project_xlsx(data))

    ws = wb["Methodology"]
    combined = "\n".join(str(c.value) for c in _all_cells(wb) if c.value and ws.title == "Methodology")
    assert "Showing 1 of 3" in combined
    assert "truncated" in combined


def test_empty_project_sheets_show_no_data_message(report_db, project_factory):
    project = project_factory()
    data = build_project_report_data(report_db, project.id, AnalyticsFilters())
    wb = _open(report_xlsx.build_project_xlsx(data))

    ws = wb["Brand Performance"]
    assert ws.cell(row=1, column=1).value == "No data available for the selected filters."


def test_ranked_sheets_capped_at_excel_top_n(report_db, project_factory, article_factory):
    project = project_factory()
    for i in range(EXCEL_TOP_N + 10):
        article_factory(project, count=1, retailer=f"Brand{i}")

    data = build_project_report_data(report_db, project.id, AnalyticsFilters())
    wb = _open(report_xlsx.build_project_xlsx(data))

    ws = wb["Brand Performance"]
    data_rows = ws.max_row - 1  # minus header
    assert data_rows == EXCEL_TOP_N


# --- MSL branding ------------------------------------------------------------


def test_executive_summary_sheet_contains_msl_logo(report_db, project_factory, article_factory):
    project = project_factory()
    article_factory(project, count=1, retailer="Auchan")

    data = build_project_report_data(report_db, project.id, AnalyticsFilters())
    wb = _open(report_xlsx.build_project_xlsx(data))

    ws = wb["Executive Summary"]
    assert len(ws._images) == 1


def test_comparison_executive_summary_sheet_contains_msl_logo(report_db, project_factory, article_factory):
    a = project_factory(name="A", quarter="2026-Q1")
    b = project_factory(name="B", quarter="2026-Q2")
    article_factory(a, count=1, retailer="Auchan")
    article_factory(b, count=1, retailer="Auchan")

    data = build_comparison_report_data(report_db, [a.id], [b.id], AnalyticsFilters())
    wb = _open(report_xlsx.build_comparison_xlsx(data))

    ws = wb["Executive Summary"]
    assert len(ws._images) == 1


def test_methodology_sheet_contains_msl_logo(report_db, project_factory, article_factory):
    project = project_factory()
    article_factory(project, count=1, retailer="Auchan")

    data = build_project_report_data(report_db, project.id, AnalyticsFilters())
    wb = _open(report_xlsx.build_project_xlsx(data))

    ws = wb["Methodology"]
    assert len(ws._images) == 1


def test_branding_does_not_break_filters_or_freeze_panes(report_db, project_factory, article_factory):
    """The logo/branded header must not obscure cells or interfere with
    freeze panes / autofilter on sheets that have them — this workbook
    must still open cleanly after the branding is added.
    """
    project = project_factory()
    for i in range(3):
        article_factory(project, count=1, retailer=f"Brand{i}")

    data = build_project_report_data(report_db, project.id, AnalyticsFilters())
    raw = report_xlsx.build_project_xlsx(data)
    wb = _open(raw)  # must open without error

    ws = wb["Brand Performance"]
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref is not None
    assert ws.cell(row=2, column=1).value in {"Brand0", "Brand1", "Brand2"}  # data row unobscured

    exec_ws = wb["Executive Summary"]
    assert exec_ws.cell(row=1, column=1).value  # brand band title present, not blank/overwritten


def test_oversized_xlsx_raises_controlled_error(report_db, project_factory, article_factory):
    project = project_factory()
    article_factory(project, count=1, retailer="Auchan")
    data = build_project_report_data(report_db, project.id, AnalyticsFilters())

    with patch("app.services.report_xlsx.MAX_XLSX_BYTES", 100):
        with pytest.raises(ReportTooLargeError):
            report_xlsx.build_project_xlsx(data)
