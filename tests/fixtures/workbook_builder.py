"""Builders for synthetic .xlsx fixtures used by the Excel parser tests.

Both builders mimic the real retail media-monitoring workbooks: a standard
15-column layout used by most retailers, and the Penny/Rewe 14-column
layout that omits "subfolder 2". Headers use the Romanian labels found in
the real files, with a banner row above the header row in the standard
workbook to exercise header-row detection.
"""

from datetime import date
from pathlib import Path

from openpyxl import Workbook

STANDARD_HEADERS = [
    "Nr crt",
    "Mediu",
    "Data",
    "Titlu",
    "Sursa",
    "Subiect",
    "Audienta",
    "AVE",
    "Tonalitate",
    "Importanta",
    "Autor",
    "Judet",
    "Audienta sursei",
    "Subfolder 1",
    "Subfolder 2",
]

PENNY_HEADERS = [h for h in STANDARD_HEADERS if h != "Subfolder 2"]


def build_standard_workbook(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Monitorizare"

    ws.cell(row=1, column=1, value="Raport monitorizare presa - Auchan Q2 2026")
    for col, header in enumerate(STANDARD_HEADERS, start=1):
        ws.cell(row=2, column=col, value=header)

    rows = [
        {
            "Nr crt": 1,
            "Mediu": "Online",
            "Data": date(2026, 4, 3),
            "Titlu": "Auchan lanseaza promotie de Paste",
            "Titlu_url": "https://mediatrust.example.com/a1",
            "Sursa": "Ziarul Financiar",
            "Sursa_url": "https://zf.ro/articol-1",
            "Subiect": "Promotii",
            "Audienta": 15000,
            "AVE": "1.234,56",
            "Tonalitate": "Pozitiv",
            "Importanta": "Ridicata",
            "Autor": "Ion Popescu",
            "Judet": "Bucuresti",
            "Audienta sursei": 500000,
            "Subfolder 1": "Retail",
            "Subfolder 2": "Promotii",
        },
        {
            "Nr crt": 2,
            "Mediu": "Print",
            "Data": "12.04.2026",
            "Titlu": "Auchan deschide un nou magazin",
            "Titlu_url": None,
            "Sursa": "Adevarul",
            "Sursa_url": "https://adevarul.ro/articol-2",
            "Subiect": "Expansiune",
            "Audienta": None,
            "AVE": 2500,
            "Tonalitate": "Neutru",
            "Importanta": "Medie",
            "Autor": None,
            "Judet": "Cluj",
            "Audienta sursei": None,
            "Subfolder 1": "Retail",
            "Subfolder 2": None,
        },
        # Exact duplicate of row 1 (same title/source/date/url) to exercise dedup.
        {
            "Nr crt": 3,
            "Mediu": "Online",
            "Data": date(2026, 4, 3),
            "Titlu": "Auchan lanseaza promotie de Paste",
            "Titlu_url": "https://mediatrust.example.com/a1-copy",
            "Sursa": "Ziarul Financiar",
            "Sursa_url": "https://zf.ro/articol-1",
            "Subiect": "Promotii",
            "Audienta": 15000,
            "AVE": "1.234,56",
            "Tonalitate": "Pozitiv",
            "Importanta": "Ridicata",
            "Autor": "Ion Popescu",
            "Judet": "Bucuresti",
            "Audienta sursei": 500000,
            "Subfolder 1": "Retail",
            "Subfolder 2": "Promotii",
        },
        # Invalid row: no title, no source.
        {
            "Nr crt": 4,
            "Mediu": "Online",
            "Data": date(2026, 4, 5),
            "Titlu": None,
            "Titlu_url": None,
            "Sursa": None,
            "Sursa_url": None,
            "Subiect": "Necunoscut",
            "Audienta": None,
            "AVE": 100,
            "Tonalitate": None,
            "Importanta": None,
            "Autor": None,
            "Judet": None,
            "Audienta sursei": None,
            "Subfolder 1": None,
            "Subfolder 2": None,
        },
        # Unparseable date and numeric value, but still a valid row (has a title).
        {
            "Nr crt": 5,
            "Mediu": "Online",
            "Data": "N/A",
            "Titlu": "Auchan in presa locala",
            "Titlu_url": None,
            "Sursa": "Presa Locala",
            "Sursa_url": None,
            "Subiect": "Comunitate",
            "Audienta": 800,
            "AVE": "abc",
            "Tonalitate": "Neutru",
            "Importanta": "Scazuta",
            "Autor": "Maria Ionescu",
            "Judet": "Iasi",
            "Audienta sursei": 12000,
            "Subfolder 1": "CSR",
            "Subfolder 2": "Comunitate",
        },
    ]

    _write_rows(ws, rows, header_row=2, headers=STANDARD_HEADERS)

    wb.save(path)
    return path


def build_penny_workbook(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Monitorizare"

    for col, header in enumerate(PENNY_HEADERS, start=1):
        ws.cell(row=1, column=col, value=header)

    rows = [
        {
            "Nr crt": 1,
            "Mediu": "Online",
            "Data": date(2026, 5, 10),
            "Titlu": "Penny Rewe extinde reteaua de magazine",
            "Titlu_url": "https://mediatrust.example.com/p1",
            "Sursa": "Ziarul Financiar",
            "Sursa_url": "https://zf.ro/articol-penny-1",
            "Subiect": "Expansiune",
            "Audienta": 9000,
            "AVE": "3.400,00",
            "Tonalitate": "Pozitiv",
            "Importanta": "Ridicata",
            "Autor": "Elena Radu",
            "Judet": "Timis",
            "Audienta sursei": 300000,
            "Subfolder 1": "Retail",
        },
        {
            "Nr crt": 2,
            "Mediu": "Print",
            "Data": date(2026, 5, 12),
            "Titlu": "Rewe raporteaza rezultate financiare",
            "Titlu_url": None,
            "Sursa": "Business Magazin",
            "Sursa_url": "https://businessmagazin.ro/articol-penny-2",
            "Subiect": "Rezultate financiare",
            "Audienta": 4200,
            "AVE": 1800.5,
            "Tonalitate": "Neutru",
            "Importanta": "Medie",
            "Autor": None,
            "Judet": "Bucuresti",
            "Audienta sursei": 150000,
            "Subfolder 1": "Financiar",
        },
    ]

    _write_rows(ws, rows, header_row=1, headers=PENNY_HEADERS)

    wb.save(path)
    return path


def _write_rows(ws, rows, header_row, headers):
    for offset, row_data in enumerate(rows, start=1):
        excel_row = header_row + offset
        for col, header in enumerate(headers, start=1):
            value = row_data.get(header)
            cell = ws.cell(row=excel_row, column=col, value=value)
            if header == "Titlu" and row_data.get("Titlu_url"):
                cell.hyperlink = row_data["Titlu_url"]
            if header == "Sursa" and row_data.get("Sursa_url"):
                cell.hyperlink = row_data["Sursa_url"]
