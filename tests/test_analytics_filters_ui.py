"""Phase D — filter persistence through navigation, export ("drill-down")
links, and clear-one/clear-all behavior in the Analytics tab.
"""

import re
from io import BytesIO

import openpyxl

from app.models.article import RetailerReviewStatus


def test_multi_brand_selection_survives_in_the_response_state(
    authenticated_client, db_session, project_factory, article_factory
):
    project = project_factory()
    article_factory(project, count=1, retailer="Auchan")
    article_factory(project, count=1, retailer="Carrefour")
    project.total_rows = 2
    db_session.commit()

    response = authenticated_client.get(
        f"/projects/{project.id}?tab=analytics&brand=Auchan&brand=Carrefour"
    )
    assert response.status_code == 200
    # Both checkboxes render pre-checked for the current selection.
    assert response.text.count('name="brand" value="Auchan" checked') == 1
    assert response.text.count('name="brand" value="Carrefour" checked') == 1


def test_source_file_selection_survives_in_the_response_state(
    authenticated_client, db_session, project_factory, article_factory, uploaded_file_factory
):
    project = project_factory()
    uploaded_file = uploaded_file_factory(project, original_filename="Auchan_Q2.xlsx")
    article_factory(project, count=1, retailer="Auchan", uploaded_file_id=uploaded_file.id)
    project.total_rows = 1
    db_session.commit()

    response = authenticated_client.get(
        f"/projects/{project.id}?tab=analytics&source_file={uploaded_file.id}"
    )
    assert response.status_code == 200
    assert f'value="{uploaded_file.id}" checked' in response.text


def test_needs_review_checkbox_state_survives(
    authenticated_client, db_session, project_factory, article_factory
):
    project = project_factory()
    article_factory(
        project, count=1, retailer="unknown",
        retailer_review_status=RetailerReviewStatus.NEEDS_REVIEW,
    )
    project.total_rows = 1
    db_session.commit()

    response = authenticated_client.get(f"/projects/{project.id}?tab=analytics&needs_review=1")
    assert response.status_code == 200
    assert 'name="needs_review" value="1" checked' in response.text


# --- export ("drill-down") links preserve the full filter state -------------


def test_export_links_preserve_multi_brand_selection(
    authenticated_client, db_session, project_factory, article_factory
):
    project = project_factory()
    article_factory(project, count=1, retailer="Auchan")
    article_factory(project, count=1, retailer="Carrefour")
    project.total_rows = 2
    db_session.commit()

    response = authenticated_client.get(
        f"/projects/{project.id}?tab=analytics&brand=Auchan&brand=Carrefour"
    )
    assert response.status_code == 200
    # Export links carry the canonical serializer's key name ("brands",
    # plural) regardless of which legacy/canonical key the request itself
    # used to select them.
    assert "brands=Auchan" in response.text
    assert "brands=Carrefour" in response.text
    assert "report.pptx?" in response.text
    assert "report.xlsx?" in response.text


def test_export_links_preserve_source_file_selection(
    authenticated_client, db_session, project_factory, article_factory, uploaded_file_factory
):
    project = project_factory()
    uploaded_file = uploaded_file_factory(project)
    article_factory(project, count=1, retailer="Auchan", uploaded_file_id=uploaded_file.id)
    project.total_rows = 1
    db_session.commit()

    response = authenticated_client.get(
        f"/projects/{project.id}?tab=analytics&source_file={uploaded_file.id}"
    )
    assert response.status_code == 200
    # Export links carry the canonical serializer's key name
    # ("source_files", plural), regardless of the request's own key.
    assert f"source_files={uploaded_file.id}" in response.text


def test_export_links_preserve_needs_review_flag(
    authenticated_client, db_session, project_factory, article_factory
):
    project = project_factory()
    article_factory(
        project, count=1, retailer="unknown",
        retailer_review_status=RetailerReviewStatus.NEEDS_REVIEW,
    )
    project.total_rows = 1
    db_session.commit()

    response = authenticated_client.get(f"/projects/{project.id}?tab=analytics&needs_review=1")
    assert response.status_code == 200
    assert "needs_review=1" in response.text


def test_export_links_preserve_all_filters_combined(
    authenticated_client, db_session, project_factory, article_factory
):
    project = project_factory()
    article_factory(project, count=1, retailer="Auchan", source="Ziarul Financiar")
    project.total_rows = 1
    db_session.commit()

    response = authenticated_client.get(
        f"/projects/{project.id}?tab=analytics&brand=Auchan&publication=Ziarul+Financiar"
    )
    assert response.status_code == 200
    assert "brands=Auchan" in response.text
    assert "publication=Ziarul" in response.text


def test_exported_kpi_matches_a_multi_brand_filtered_analytics_call(
    authenticated_client, internal_headers, db_session, project_factory, article_factory
):
    """The same consistency guarantee Phase 6C already established for a
    single-brand filter, now proven for the new multi-brand shape."""
    project = project_factory()
    article_factory(project, count=2, retailer="Auchan")
    article_factory(project, count=1, retailer="Carrefour")
    article_factory(project, count=1, retailer="Lidl")
    project.total_rows = 4
    db_session.commit()

    api_response = authenticated_client.get(
        f"/api/internal/projects/{project.id}/analytics?brand=Auchan&brand=Carrefour",
        headers=internal_headers,
    )
    api_unique_valid = api_response.json()["kpis"]["unique_valid_articles"]
    assert api_unique_valid == 3

    xlsx_response = authenticated_client.get(
        f"/projects/{project.id}/report.xlsx?brand=Auchan&brand=Carrefour"
    )
    wb = openpyxl.load_workbook(BytesIO(xlsx_response.content))
    ws = wb["KPI Summary"]
    row_values = {
        ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)
    }
    assert row_values["Unique valid articles"] == api_unique_valid


# --- clear-one and clear-all -------------------------------------------------


def test_clear_all_link_is_present_and_bare(
    authenticated_client, db_session, project_factory, article_factory
):
    project = project_factory()
    article_factory(project, count=1, retailer="Auchan")
    project.total_rows = 1
    db_session.commit()

    response = authenticated_client.get(f"/projects/{project.id}?tab=analytics&brand=Auchan")
    assert response.status_code == 200
    assert f'href="/projects/{project.id}?tab=analytics"' in response.text


def test_clear_one_filter_chip_removes_only_that_dimension(
    authenticated_client, db_session, project_factory, article_factory
):
    project = project_factory()
    article_factory(project, count=1, retailer="Auchan", source="Ziarul Financiar")
    project.total_rows = 1
    db_session.commit()

    response = authenticated_client.get(
        f"/projects/{project.id}?tab=analytics&brand=Auchan&publication=Ziarul+Financiar"
    )
    assert response.status_code == 200
    assert "Active filters" in response.text
    assert "Brand: Auchan" in response.text
    assert "Publication: Ziarul Financiar" in response.text

    # The brand chip's clear link must drop only `brand`, keeping
    # `publication` (and `tab=analytics`) intact.
    match = re.search(r'href="([^"]*)"[^>]*>\s*Brand: Auchan', response.text)
    assert match is not None
    clear_url = match.group(1).replace("&amp;", "&")
    assert "brand=" not in clear_url
    assert "publication=Ziarul" in clear_url
    assert "tab=analytics" in clear_url

    follow_up = authenticated_client.get(clear_url)
    assert follow_up.status_code == 200
    assert "Brand: Auchan" not in follow_up.text
    assert "Publication: Ziarul Financiar" in follow_up.text


def test_no_active_filters_shows_no_chip_bar(
    authenticated_client, db_session, project_factory, article_factory
):
    project = project_factory()
    article_factory(project, count=1, retailer="Auchan")
    project.total_rows = 1
    db_session.commit()

    response = authenticated_client.get(f"/projects/{project.id}?tab=analytics")
    assert response.status_code == 200
    assert "Active filters" not in response.text
